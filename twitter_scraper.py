import sys
import os
import time
from fake_headers import Headers
import requests
from bs4 import BeautifulSoup
from scroller import Scroller
from openpyxl import Workbook
import pandas as pd
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.common.exceptions import (
    NoSuchElementException,
    StaleElementReferenceException,
    WebDriverException,
)

from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

from webdriver_manager.chrome import ChromeDriverManager

TWITTER_LOGIN_URL = "https://x.com/i/flow/login"
WAITTING_TIME_SECOND = 10
VERIRY_HTTPS_URL = "https://2fa.wiki"
PASSWORD_KEY = "OMW6JIO3FNTWJOFG"


class TwitterScraper:
    def __init__(self, username, password, account):
        print("Initializing Twitter Scraper...")

        self.username = username
        self.password = password
        self.account = account

        self.search_account_name = None

        self.driver = self._get_driver()
        self.scroller = Scroller(self.driver)

    def _get_driver(self):
        print("Setup WebDriver...")
        header = Headers().generate()["User-Agent"]

        # browser_option = EdgeOptions()
        browser_option = ChromeOptions()

        browser_option.add_argument("--no-sandbox")
        browser_option.add_argument("--disable-dev-shm-usage")
        browser_option.add_argument("--ignore-certificate-errors")
        browser_option.add_argument("--disable-gpu")
        browser_option.add_argument("--log-level=3")
        browser_option.add_argument("--disable-notifications")
        browser_option.add_argument("--disable-popup-blocking")
        browser_option.add_argument("--user-agent={}".format(header))

        try:
            print("Initializing ChromeDriver...")
            driver = webdriver.Chrome(
                options=browser_option,
            )

            print("WebDriver Setup Complete")
            return driver
        except WebDriverException:
            try:
                print("Downloading ChromeDriver...")
                chromedriver_path = ChromeDriverManager().install()
                chrome_service = ChromeService(
                    executable_path=chromedriver_path)

                print("Initializing ChromeDriver...")
                driver = webdriver.Chrome(
                    service=chrome_service,
                    options=browser_option,
                )

                print("WebDriver Setup Complete")
                return driver
            except Exception as e:
                print(f"Error setting up WebDriver: {e}")
                sys.exit(1)

    def login(self, verify_code):
        """
        login twitter
        """
        print()
        print("Logging in to Twitter...")

        try:
            self.driver.maximize_window()
            self.driver.get(TWITTER_LOGIN_URL)
            time.sleep(3)

            self._input_username()
            self._input_password()
            self._input_unusual_activity(verify_code)

            cookies = self.driver.get_cookies()
            auth_token = None
            for cookie in cookies:
                if cookie["name"] == "auth_token":
                    auth_token = cookie["value"]
                    break

            if auth_token is None:
                raise ValueError(
                    """This may be due to the following:

- Internet connection is unstable
- Username is incorrect
- Password is incorrect
"""
                )

            print()
            print("Login Successful")
            print()

        except Exception as e:
            print()
            print(f"Login Failed: {e}")
            sys.exit(1)

    def _input_username(self):
        """
        input username
        """
        input_attempt = 0

        while True:
            try:
                username = self.driver.find_element(By.XPATH, "//input[@autocomplete='username']")
                username.send_keys(self.username)
                time.sleep(WAITTING_TIME_SECOND)
                username.send_keys(Keys.RETURN)
                time.sleep(WAITTING_TIME_SECOND // 2)
                break
            except NoSuchElementException:
                input_attempt += 1
                if input_attempt >= 3:
                    print()
                    print(
                        """There was an error inputting the username.

It may be due to the following:
- Internet connection is unstable
- Username is incorrect
- Twitter is experiencing unusual activity"""
                    )
                    self.driver.quit()
                    sys.exit(1)
                else:
                    print("Re-attempting to input username...")
                    time.sleep(2)

    def _input_password(self):
        input_attempt = 0

        while True:
            try:
                password = self.driver.find_element(By.XPATH, "//input[@autocomplete='current-password']")

                password.send_keys(self.password)
                time.sleep(WAITTING_TIME_SECOND)
                password.send_keys(Keys.RETURN)
                time.sleep(WAITTING_TIME_SECOND // 2)
                break
            except NoSuchElementException:
                input_attempt += 1
                if input_attempt >= 3:
                    print()
                    print(
                        """There was an error inputting the password.

It may be due to the following:
- Internet connection is unstable
- Password is incorrect
- Twitter is experiencing unusual activity"""
                    )
                    self.driver.quit()
                    sys.exit(1)
                else:
                    print("Re-attempting to input password...")
                    time.sleep(2)
    
    def _input_verify_code(self):
        pass
    
    def _input_unusual_activity(self, verify_code):
        input_attempt = 0

        while True:
            try:
                unusual_activity = self.driver.find_element(By.XPATH, "//input[@data-testid='ocfEnterTextTextInput']")
                unusual_activity.send_keys(verify_code)
                time.sleep(WAITTING_TIME_SECOND)
                unusual_activity.send_keys(Keys.RETURN)
                time.sleep(WAITTING_TIME_SECOND // 2)
                break
            except NoSuchElementException:
                input_attempt += 1
                if input_attempt >= 3:
                    break

    def go_to_home(self):
        """
        return to twitter home page
        """
        self.driver.get("https://x.com/home")
        time.sleep(3)
        pass

    def get_search_account_name(self):
        return self.search_account_name

    def back(self):
        """
        Go back to the previous page.
        """
        self.driver.back()
        time.sleep(WAITTING_TIME_SECOND // 3)

    def quit(self):
        """
        quit chrome
        """
        self.driver.quit()

    def search(self, query_username, intetractive=True):
        search_element = self.driver.find_element(By.XPATH, "//input[@placeholder='Search']")
        search_element.send_keys(query_username)
        time.sleep(WAITTING_TIME_SECOND)
        search_element.send_keys(Keys.RETURN)
        time.sleep(WAITTING_TIME_SECOND)

        self._go_to_people(query_username)
        not_lock = self._go_to_query_user(query_username, intetractive)

        if not_lock:
            self._go_to_following(self.search_account_name)
            self._go_to_followers(self.search_account_name)
            self.back()
            self.back()
            self.back()
            self.back()
            self.back()
        else:
            self.back()
            self.back()
            self.back()

    def _go_to_people(self, search_user):
        """
        go to People page
        """
        span_elements = self.driver.find_elements(By.TAG_NAME, "span")
        people_span = None
        for span in span_elements:
            if span.text == 'People':
                people_span = span
                break
        if people_span:
            people_span.click()
            time.sleep(WAITTING_TIME_SECOND)
        else:
            print("Not Found People Element.")
            sys.exit(1)
        
    def _go_to_query_user(self, search_user, intetractive):
        try:
            span_elements = self.driver.find_elements(By.TAG_NAME, "span")
            user_spans = list()
            for span in span_elements:
                if span.text == search_user or search_user in span.text:
                    user_spans.append(span)

            if intetractive:
                print("all users queryed.")
                for i in range(len(user_spans)):
                    print(f"{i+1}. {user_spans[i].text}")
                print("Enter number to select one user to query: ")
                number = int(sys.stdin.readline().strip())
                
                # 点击
                if number < 1 or number > len(user_spans):
                    print("Enter true number to select one user to query: ")
                    number = int(sys.stdin.readline().strip())
            
                user_spans[number-1].click()
            else:
                user_spans[0].click()
            
            time.sleep(WAITTING_TIME_SECOND // 4)
            try:
                self.search_account_name = self.driver.find_element(By.CSS_SELECTOR, 'div[data-testid="UserName"] div.css-146c3p1[style*="color: rgb(113, 118, 123)"] span.css-1jxf684').text[1:]
            except NoSuchElementException:
                self.search_account_name = search_user

            time.sleep(WAITTING_TIME_SECOND)

            try:
                self.driver.find_element(By.XPATH, "//svg[@data-testid='icon-lock']")
                return False
            except NoSuchElementException:
                return True
            

        except NoSuchElementException as e:
            print(f"Not Found User Name {search_user} to search.")
            sys.exit(1)
        
        return True

    def _go_to_following(self, search_user):
        """
        go to following page
        """
        span_elements = self.driver.find_elements(By.TAG_NAME, "span")
        following_span = None
        for span in span_elements:
            if span.text == 'Following':
                following_span = span
                break
        if following_span:
            following_span.click()
            time.sleep(WAITTING_TIME_SECOND // 2)
            self._get_following_or_follower_users(search_user, following=True)
            time.sleep(WAITTING_TIME_SECOND)
            
        else:
            print("Not Found Following Element.")
            sys.exit(1)

    def _get_following_or_follower_users(self, search_user, following=True):
        """
        获取用户的关注情况, following为真表示获取following, 反之表示获取follower
        """
        # 获取页面初始的滚动高度
        last_height = self.driver.execute_script("return document.body.scrollHeight")
        users_account_name_set = set()
        while True:
            # 每次滚动一个屏幕高度
            self.driver.execute_script("window.scrollBy(0, window.innerHeight);")

            # 等待页面加载新内容
            time.sleep(WAITTING_TIME_SECOND)

            try:
                following_user_div_elements = self.driver.find_elements(By.XPATH, "//div[@data-testid='cellInnerDiv']")
                
                users_list = list()
                for element in following_user_div_elements:
                    try:
                        user_name = ''
                        user_name = element.find_element(By.CSS_SELECTOR, 'div.css-146c3p1.r-bcqeeo.r-1ttztb7.r-qvutc0.r-37j5jr.r-a023e6.r-rjixqe.r-b88u0q.r-1awozwy.r-6koalj.r-1udh08x.r-3s2u2q span.css-1jxf684.r-bcqeeo.r-1ttztb7.r-qvutc0.r-poiln3').text
                        print(user_name)
                    except NoSuchElementException:
                        pass

                    try:
                        accout_name = ''
                        accout_name = element.find_element(By.CSS_SELECTOR, 'div.css-146c3p1.r-dnmrzs.r-1udh08x.r-1udbk01.r-3s2u2q.r-bcqeeo.r-1ttztb7.r-qvutc0.r-37j5jr.r-a023e6.r-rjixqe.r-16dba41.r-18u37iz.r-1wvb978 span.css-1jxf684.r-bcqeeo.r-1ttztb7.r-qvutc0.r-poiln3').text
                        print(accout_name)
                    except NoSuchElementException:
                        pass

                    try:
                        description = ''
                        description = element.find_element(By.CSS_SELECTOR, 'div.css-146c3p1.r-bcqeeo.r-1ttztb7.r-qvutc0.r-37j5jr.r-a023e6.r-rjixqe.r-16dba41.r-1h8ys4a.r-1jeg54m').text
                        print(description)
                    except NoSuchElementException:
                        pass
                    
                    if accout_name not in users_account_name_set:
                        users_account_name_set.add(accout_name)
                        users_list.append(tuple([user_name, accout_name, description]))

                self._write_to_excel(f'{search_user}.xlsx', users_list, following)
            except NoSuchElementException as e:
                print(f"Not Found users.")
                # sys.exit(1)

            # 获取当前页面的滚动高度
            new_height = self.driver.execute_script("return document.body.scrollHeight")

            # 判断是否已经滚动到底部
            if new_height == last_height:
                break
            last_height = new_height
        
        self.driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(WAITTING_TIME_SECOND // 2)

    def _write_to_excel(self, excel_file, users_list, following=True):
        current_dir = os.getcwd()
        file_path = os.path.join(current_dir, excel_file)

        if not os.path.exists(file_path):
            wb = Workbook()
            ws1 = wb.active
            ws1.title = 'following'
            wb.create_sheet('follower')
            wb.save(file_path)

        wb = load_workbook(file_path)
        
        # 将数据写入文件
        df = pd.DataFrame(users_list, columns=['username', 'accountname', 'description'])

        sheet_name = 'following' if following else 'follower'

        ws = wb[sheet_name]

        # 如果工作表是空的，写入表头
        if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
            for col_num, col_name in enumerate(df.columns, start=1):
                ws.cell(row=1, column=col_num, value=col_name)

        # 从第二行开始写入数据
        start_row = ws.max_row + 1
        for row_num, row_data in df.iterrows():
            for col_num, value in enumerate(row_data, start=1):
                ws.cell(row=start_row + row_num, column=col_num, value=value)

        # 保存修改后的 Excel 文件
        wb.save(file_path)

    def _go_to_followers(self, search_user):
        """
        go to followers page
        """
        span_elements = self.driver.find_elements(By.TAG_NAME, "span")
        followers_span = None
        for span in span_elements:
            if span.text == 'Followers':
                followers_span = span
                break
        if followers_span:
            followers_span.click()
            time.sleep(WAITTING_TIME_SECOND // 2)
            self._get_following_or_follower_users(search_user, following=False)
            time.sleep(WAITTING_TIME_SECOND)
        else:
            print("Not Found Followers Element.")
            sys.exit(1)

    def get_verify_code(self):
        """
        return verify code
        """
        # 设置HTTP头信息
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive'
        }

        # 发送GET请求
        response = requests.get(VERIRY_HTTPS_URL, headers=headers)

        soup = BeautifulSoup(response.text, 'lxml')
        secret_tag = soup.find('span', attrs={'class': 'secret'})
        verify_code = ''
        if secret_tag:
            verify_code = secret_tag.h3.text
        else:
            self.driver.get(VERIRY_HTTPS_URL)
            time.sleep(WAITTING_TIME_SECOND // 3)
            
            # 点击添加密钥按钮
            add_button = self.driver.find_element(By.XPATH, "//a[@id='addButton']")
            add_button.click()
            time.sleep(WAITTING_TIME_SECOND // 3)

            # 输入名称
            account = self.driver.find_element(By.XPATH, "//input[@name='keyAccount']")
            account.send_keys('aaa')
            time.sleep(WAITTING_TIME_SECOND // 3)

            # 输入密钥
            account = self.driver.find_element(By.XPATH, "//input[@name='keySecret']")
            account.send_keys(PASSWORD_KEY)
            time.sleep(WAITTING_TIME_SECOND // 3)

            # 点击生成
            add_key_button = self.driver.find_element(By.XPATH, "//a[@id='addKeyButton']")
            add_key_button.click()
            time.sleep(WAITTING_TIME_SECOND // 3)

            # 获取代码
            element = self.driver.find_element(By.XPATH, "//span[@class='secret']")
            h3_element = element.find_element(By.TAG_NAME, "h3")
            verify_code = h3_element.text
        
        return verify_code




